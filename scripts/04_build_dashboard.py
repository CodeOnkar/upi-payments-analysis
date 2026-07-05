"""
Step 4: Build an Excel dashboard workbook from the cleaned data.

Input:  data/upi_cleaned.csv
Output: outputs/UPI_Dashboard.xlsx

What this builds:
  - Sheet "Data": the cleaned monthly data (source for all charts)
  - Sheet "Yearly": annual rollups + linear trend forecast for 2022-2023
  - Sheet "Dashboard": KPI cards, line chart (volume/value trends),
    bar chart (YoY growth %)

Forecast method (kept deliberately simple, explained in the sheet itself):
  Ordinary least-squares straight-line fit on FULL calendar years only
  (2017-2021; 2016 excluded because it's a partial 9-month year and would
  distort the slope). The line is extended to 2022 and 2023. This is a
  naive baseline forecast -- it assumes the past *rate* of increase
  (a constant amount added per year) continues, so it will systematically
  UNDERSTATE growth if the real trend is accelerating (which annual value
  more than doubling in 2021 suggests it is). We say this explicitly in
  the workbook rather than presenting the forecast as authoritative.
"""

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

df = pd.read_csv("data/upi_cleaned.csv", parse_dates=["date"])
df["year"] = df["date"].dt.year

yearly = (
    df.groupby("year")
    .agg(
        total_volume_mn=("volume_mn", "sum"),
        total_value_cr=("value_cr", "sum"),
        banks_live_year_end=("banks_live", "max"),
    )
    .reset_index()
)

# --- Linear trend forecast (2017-2021 only; 2016 is a partial year) ---
fit = yearly[yearly["year"] >= 2017]
slope_val, intercept_val = np.polyfit(fit["year"], fit["total_value_cr"], 1)
slope_vol, intercept_vol = np.polyfit(fit["year"], fit["total_volume_mn"], 1)

forecast_rows = []
for year in [2022, 2023]:
    forecast_rows.append({
        "year": year,
        "total_volume_mn": slope_vol * year + intercept_vol,
        "total_value_cr": slope_val * year + intercept_val,
        "banks_live_year_end": np.nan,
        "is_forecast": True,
    })
yearly["is_forecast"] = False
yearly_full = pd.concat([yearly, pd.DataFrame(forecast_rows)], ignore_index=True)

# ============================================================
# Build workbook
# ============================================================
wb = Workbook()

HEADER_FILL = PatternFill("solid", fgColor="1F6FEB")
HEADER_FONT = Font(color="FFFFFF", bold=True)
FORECAST_FILL = PatternFill("solid", fgColor="FFF3CD")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_df(ws, dataframe, start_row=1, highlight_col=None):
    for j, col in enumerate(dataframe.columns, start=1):
        cell = ws.cell(row=start_row, column=j, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
    for i, row in enumerate(dataframe.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = BORDER
            if highlight_col is not None and getattr(row, highlight_col, False):
                cell.fill = FORECAST_FILL
    for j, col in enumerate(dataframe.columns, start=1):
        ws.column_dimensions[get_column_letter(j)].width = max(14, len(str(col)) + 2)


# --- Sheet: Data (monthly cleaned data) ---
ws_data = wb.active
ws_data.title = "Data"
data_cols = ["date", "banks_live", "volume_mn", "value_cr",
             "avg_ticket_size_inr", "mom_growth_pct", "yoy_growth_pct"]
df_out = df[data_cols].copy()
df_out["date"] = df_out["date"].dt.strftime("%Y-%m-%d")
write_df(ws_data, df_out)

# --- Sheet: Yearly (rollups + forecast) ---
ws_year = wb.create_sheet("Yearly")
yearly_cols = ["year", "total_volume_mn", "total_value_cr", "banks_live_year_end", "is_forecast"]
write_df(ws_year, yearly_full[yearly_cols], highlight_col="is_forecast")

note_row = len(yearly_full) + 3
ws_year.cell(row=note_row, column=1,
             value="Forecast note: 2022-2023 figures (highlighted) are a simple linear "
                   "trend line fit on FULL years 2017-2021 only (2016 excluded: partial "
                   "9-month year). This is a naive baseline -- it assumes a constant "
                   "yearly increase, so it likely UNDERSTATES 2022-23 value given that "
                   "annual value more than doubled in 2021 alone (accelerating growth).")
ws_year.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)
ws_year.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
ws_year.row_dimensions[note_row].height = 60

# ============================================================
# Sheet: Dashboard
# ============================================================
ws_dash = wb.create_sheet("Dashboard", 0)  # make it the first/active sheet
ws_dash.sheet_view.showGridLines = False

TITLE_FONT = Font(size=18, bold=True, color="1F2937")
ws_dash["B2"] = "UPI Payments India -- Dashboard (2016-2021)"
ws_dash["B2"].font = TITLE_FONT

# --- KPI cards ---
last_row = df.iloc[-1]
total_value_all_time = df["value_cr"].sum()
total_volume_all_time = df["volume_mn"].sum()
peak_growth_row = df[(df["date"] >= "2018-01-01") & df["mom_growth_pct"].notna()] \
    .sort_values("mom_growth_pct", ascending=False).iloc[0]

kpis = [
    ("Total Transactions (Volume)", f"{total_volume_all_time:,.0f} Mn"),
    ("Total Transaction Value", f"Rs {total_value_all_time:,.0f} Cr"),
    ("Peak MoM Growth Month", f"{peak_growth_row['date'].strftime('%b %Y')} (+{peak_growth_row['mom_growth_pct']:.1f}%)"),
    ("Banks Live on UPI (Dec 2021)", f"{int(last_row['banks_live'])}"),
]

KPI_FILL = PatternFill("solid", fgColor="EFF6FF")
KPI_LABEL_FONT = Font(size=10, color="6B7280")
KPI_VALUE_FONT = Font(size=14, bold=True, color="1F2937")

kpi_start_col = 2  # column B
for i, (label, value) in enumerate(kpis):
    col = kpi_start_col + i * 3
    row = 4
    for r in range(row, row + 4):
        for c in range(col, col + 2):
            ws_dash.cell(row=r, column=c).fill = KPI_FILL
            ws_dash.cell(row=r, column=c).border = BORDER
    label_cell = ws_dash.cell(row=row, column=col, value=label)
    label_cell.font = KPI_LABEL_FONT
    label_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws_dash.merge_cells(start_row=row, start_column=col, end_row=row + 1, end_column=col + 1)
    value_cell = ws_dash.cell(row=row + 2, column=col, value=value)
    value_cell.font = KPI_VALUE_FONT
    ws_dash.merge_cells(start_row=row + 2, start_column=col, end_row=row + 3, end_column=col + 1)

for i in range(len(kpis) * 3):
    ws_dash.column_dimensions[get_column_letter(kpi_start_col + i)].width = 15

# --- Hidden helper data on Dashboard sheet (charts need a Reference range) ---
# We reference the Data and Yearly sheets directly instead of duplicating data.

# Line chart: Volume & Value trend (dual chart, two separate charts to avoid
# misleading dual-axis scaling -- same reasoning as the matplotlib EDA step)
chart_vol = LineChart()
chart_vol.title = "Monthly Transaction Volume (Mn)"
chart_vol.y_axis.title = "Volume (Mn)"
chart_vol.x_axis.title = "Month"
chart_vol.height = 8
chart_vol.width = 16
data_ref = Reference(ws_data, min_col=3, min_row=1, max_row=ws_data.max_row)
cats_ref = Reference(ws_data, min_col=1, min_row=2, max_row=ws_data.max_row)
chart_vol.add_data(data_ref, titles_from_data=True)
chart_vol.set_categories(cats_ref)
ws_dash.add_chart(chart_vol, "B10")

chart_val = LineChart()
chart_val.title = "Monthly Transaction Value (Cr INR)"
chart_val.y_axis.title = "Value (Cr INR)"
chart_val.x_axis.title = "Month"
chart_val.height = 8
chart_val.width = 16
data_ref2 = Reference(ws_data, min_col=4, min_row=1, max_row=ws_data.max_row)
chart_val.add_data(data_ref2, titles_from_data=True)
chart_val.set_categories(cats_ref)
ws_dash.add_chart(chart_val, "K10")

# Bar chart: YoY growth % (2018 onward, consistent with EDA step's reasoning
# about small-base distortion in 2016-17)
yoy_start_row = df[df["date"] >= "2018-01-01"].index.min() + 2  # +2: header row + 0-index offset
chart_yoy = BarChart()
chart_yoy.title = "YoY Growth % (2018 onward)"
chart_yoy.y_axis.title = "YoY Growth (%)"
chart_yoy.x_axis.title = "Month"
chart_yoy.height = 8
chart_yoy.width = 16
data_ref3 = Reference(ws_data, min_col=7, min_row=yoy_start_row - 1, max_row=ws_data.max_row)
cats_ref3 = Reference(ws_data, min_col=1, min_row=yoy_start_row, max_row=ws_data.max_row)
chart_yoy.add_data(data_ref3, titles_from_data=True)
chart_yoy.set_categories(cats_ref3)
ws_dash.add_chart(chart_yoy, "B28")

# Bar chart: Forecast (Yearly sheet: total_value_cr including 2022-23 forecast)
chart_forecast = BarChart()
chart_forecast.title = "Total Annual Value: Actual (2016-2021) vs Linear Forecast (2022-2023)"
chart_forecast.y_axis.title = "Total Value (Cr INR)"
chart_forecast.x_axis.title = "Year"
chart_forecast.height = 8
chart_forecast.width = 16
fdata_ref = Reference(ws_year, min_col=3, min_row=1, max_row=len(yearly_full) + 1)
fcats_ref = Reference(ws_year, min_col=1, min_row=2, max_row=len(yearly_full) + 1)
chart_forecast.add_data(fdata_ref, titles_from_data=True)
chart_forecast.set_categories(fcats_ref)
ws_dash.add_chart(chart_forecast, "K28")

wb.save("outputs/UPI_Dashboard.xlsx")
print("Saved outputs/UPI_Dashboard.xlsx")
print(f"KPIs: {kpis}")
